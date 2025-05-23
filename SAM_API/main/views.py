from django.shortcuts import render
from rest_framework.views import APIView
from rest_framework.generics import ListAPIView
from rest_framework.response import Response
from rest_framework import status
from .serializers import *
from rest_framework.permissions import AllowAny
from .models import TeacherUsersStats, TeacherTopic
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from django.utils.timezone import is_aware
from django.http import HttpResponse
import io
from openpyxl.utils import get_column_letter
from rest_framework.generics import RetrieveUpdateAPIView

__all__ = [
    'TeacherUsersStatsView',
    'ExportToExcelView',
    'TeacherEditView',
    'TopicsView',
    'TopicedTeachersView'
]


class TopicsView(ListAPIView):
    permission_classes = [AllowAny]
    queryset = TeacherTopic.objects.all()
    serializer_class = TopicsSerializer

class TeacherUsersStatsView(APIView):
    serializer_class = TeacherUsersStatsSerializer
    permission_classes = [AllowAny]
    def get(self, request):
        stats = TeacherUsersStats.objects.all()
        serializer = self.serializer_class(stats, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

class TeacherEditView(RetrieveUpdateAPIView):
    serializer_class = TeacherEditSerializer
    permission_classes = [AllowAny]
    queryset = TeacherUsersStats.objects.all()
    lookup_field = 'id'

    def get(self, request, *args, **kwargs):
        if self.kwargs.get('id') is None:
            return Response(
                {"detail": "ID kiritilishi shart."},
                status=status.HTTP_400_BAD_REQUEST
            )
        return super().get(request, *args, **kwargs)


class ExportToExcelView(APIView):
    permission_classes = [AllowAny]

    def get(self, request):
        queryset = TeacherUsersStats.objects.all()
        variants = ["Juda yaxshi", "Yaxshi", "O‘rtacha", "Past", "Yomon"]

        wb = Workbook()
        ws = wb.active
        ws.title = "TeacherStats"

        ws.cell(row=1, column=1, value="F.I.O")
        ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)

        col = 2
        for i in range(1, 7):
            ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + 4)
            ws.cell(row=1, column=col, value=f"{i}-mezon")
           
            for j, variant in enumerate(variants):
                ws.cell(row=2, column=col + j, value=variant)
            col += 5

        row = 3
        for obj in queryset:
            ws.cell(row=row, column=1, value=obj.full_name)
            col = 2
            for i in range(1, 7):
                ws.cell(row=row, column=col + 0, value=getattr(obj, f"m{i}_juda_yaxshi", 0))
                ws.cell(row=row, column=col + 1, value=getattr(obj, f"m{i}_yaxshi", 0))
                ws.cell(row=row, column=col + 2, value=getattr(obj, f"m{i}_ortacha", 0))
                ws.cell(row=row, column=col + 3, value=getattr(obj, f"m{i}_past", 0))
                ws.cell(row=row, column=col + 4, value=getattr(obj, f"m{i}_yomon", 0))
                col += 5
            row += 1

        for row_cells in ws.iter_rows():
            for cell in row_cells:
                cell.alignment = Alignment(horizontal="center", vertical="center")

        max_lengths = {}
        for row_cells in ws.iter_rows():
            for cell in row_cells:
                if cell.value:
                    col_letter = get_column_letter(cell.column)
                    max_len = len(str(cell.value))
                    if col_letter not in max_lengths or max_len > max_lengths[col_letter]:
                        max_lengths[col_letter] = max_len

        for col_letter, max_len in max_lengths.items():
            ws.column_dimensions[col_letter].width = max_len * 1.2 + 2

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=teacher_users_stats.xlsx'

        return response        
class TopicedTeachersView(ListAPIView):
    queryset = TeacherUsersStats.objects.prefetch_related('topics').all()
    permission_classes = [AllowAny]
    serializer_class = TopicedTeachersSerializer
